import requests
import re
import time
import random
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from gevent import monkey
monkey.patch_all()
import gevent
from gevent.queue import Queue
import openpyxl

# 创建ua池
def ua_pond():
    ua = UserAgent(verify_ssl=False)
    return ua.random

# 创建ip代理池
def proxy_pond():
    proxies = [
        {'http':'112.87.78.171:4216'},{'http':'222.220.154.102:4267'},
        {'http':'180.114.147.101:4283'},{'http':'182.246.158.175:4263'},
        {'http':'114.229.10.238:4207'}
        ]
    return random.sample(proxies,1)[0]

# 创建excel存储数据
def my_url():
    # 创建工作簿
    wb = openpyxl.Workbook()
    # 添加工作表
    sheet = wb.active
    # 给工作表命名
    sheet.title = 'data'
    sheet['A1'] = 'topic'
    sheet['B1'] = 'title'
    sheet['C1'] = 'url'
    wb.create_sheet(title='id',index=0)
    # 保存工作簿并命名
    wb.save('zhihu.xlsx')

# 写入存储的id
def add_id(s):
    # 打开工作簿
    wb = openpyxl.load_workbook('zhihu.xlsx')
    # 获取工作表
    sheet = wb['id']
    sheet.append(s)
    wb.save('zhihu.xlsx')

# 写入存储的数据
def add_data(s):
    # 打开工作簿
    wb = openpyxl.load_workbook('zhihu.xlsx')
    # 获取工作表
    sheet = wb['data']
    sheet.append(s)
    wb.save('zhihu.xlsx')

def rea_id():
    # 打开工作簿
    wb = openpyxl.load_workbook('zhihu.xlsx')
    # 获取工作表
    sheet = wb['id']
    rows=sheet.max_row   #获取行数
    cols=sheet.max_column    #获取列数
    id = []
    for cols in range(1,cols+1):
        Data=sheet.cell(row=rows,column=cols).value #class str
        id.append(Data)
    return id

# 定义生产者，爬取所有话题的url
def produce():
    id = []
    true_id = []
    basic_url = 'https://www.zhihu.com/topics'
    header = {'User-Agent':ua_pond()}
    # 多次爬取后这里使用的ip可能被封
    res = requests.get(basic_url,headers=header,proxies=proxy_pond())
    soup = BeautifulSoup(res.text,'html.parser')
    f = soup.find(class_='zm-topic-cat-item')
    f_id = f['data-id']
    f_name = f.text
    #构造父话题url
    url = basic_url+'#'+f_name
    header = {'User-Agent': ua_pond()}
    res1 = requests.get(url,headers=header,proxies=proxy_pond())
    #print(res1.status_code)
    soup1 = BeautifulSoup(res1.text,'html.parser')
    s_topic = soup1.find(class_='zh-general-list clearfix').find_all(class_='blk')
    # 子话题
    for s in s_topic:
        s_id = re.findall('.*?(\d.*\d)',s.find('a')['href'])
        #s_name = s.find('a')['alt']
        id.append(s_id[0])
    # 爬取更多子话题
    headers1 = {
        'origin':'https://www.zhihu.com',
        'referer':'https://www.zhihu.com/topics',
        'user-agent':ua_pond()
        }
    data = {
        'method':'next',
        'params':'{"topic_id":'+f_id+',"offset":0,"hash_id":""}'
        }
    res2 = requests.post('https://www.zhihu.com/node/TopicsPlazzaListV2',headers=headers1,data=data,proxies=proxy_pond())
    soup2 = res2.json()
    m = 0
    for s in soup2['msg']:
        s_id = re.findall('href="/topic/(\d.*\d)">',s)
        #s_name = re.findall('alt="(.*)"',s)
        id.append(s_id[0])
        m += 1
    # 如果还有子话题，继续爬取；否则，停止。
    while len(str(soup1.find(class_='zg-btn-white zu-button-more'))) != 0:
        data = {
            'method':'next',
            'params':'{"topic_id":'+f_id+',"offset":'+str(m)+',"hash_id":""}'
        }
        res2 = requests.post('https://www.zhihu.com/node/TopicsPlazzaListV2',headers=headers1,data=data,proxies=proxy_pond())
        if res2.json()['msg'] == []:
            break
        else:
            for s in res2.json()['msg']:
                s_id = re.findall('href="/topic/(\d.*\d)">',s)
                # s_name = re.findall('alt="(.*)"',s)
                id.append(s_id[0])
                m += 1
                print(m)
    # 爬取每个子话题里的子话题
    t_id = id
    for i in id:
        url = 'https://www.zhihu.com/topic/'+str(i)+'/hot'
        headers = {'User-Agent':ua_pond()}
        r = requests.get(url,headers=headers,proxies=proxy_pond())
        res = BeautifulSoup(r.text,'html.parser')
        soup = res.find_all(class_='TopicLink TopicTag')
        for i in soup:
            d = re.findall('.*?(\d.*\d)',i['href'])
            t_id.append(d[0])
    true_id = set(t_id) # class set {'','',''}，像字典
    r_id = []
    for i in true_id:
        r_id.append(i)
    add_id(r_id)

# 定义消费者，爬取所有话题里的问题
def consumer():
    # 读取话题id
    ids = rea_id()
    # 存储url
    work = Queue()
    for id in ids:
        # 构造所有话题的url
        url = 'https://www.zhihu.com/api/v4/topics/'+id+'/feeds/top_activity?include=data%5B%3F%28target.type%3Dtopic_sticky_module%29%5D.target.data%5B%3F%28target.type%3Danswer%29%5D.target.content%2Crelationship.is_authorized%2Cis_author%2Cvoting%2Cis_thanked%2Cis_nothelp%3Bdata%5B%3F%28target.type%3Dtopic_sticky_module%29%5D.target.data%5B%3F%28target.type%3Danswer%29%5D.target.is_normal%2Ccomment_count%2Cvoteup_count%2Ccontent%2Crelevant_info%2Cexcerpt.author.badge%5B%3F%28type%3Dbest_answerer%29%5D.topics%3Bdata%5B%3F%28target.type%3Dtopic_sticky_module%29%5D.target.data%5B%3F%28target.type%3Darticle%29%5D.target.content%2Cvoteup_count%2Ccomment_count%2Cvoting%2Cauthor.badge%5B%3F%28type%3Dbest_answerer%29%5D.topics%3Bdata%5B%3F%28target.type%3Dtopic_sticky_module%29%5D.target.data%5B%3F%28target.type%3Dpeople%29%5D.target.answer_count%2Carticles_count%2Cgender%2Cfollower_count%2Cis_followed%2Cis_following%2Cbadge%5B%3F%28type%3Dbest_answerer%29%5D.topics%3Bdata%5B%3F%28target.type%3Danswer%29%5D.target.annotation_detail%2Ccontent%2Chermes_label%2Cis_labeled%2Crelationship.is_authorized%2Cis_author%2Cvoting%2Cis_thanked%2Cis_nothelp%3Bdata%5B%3F%28target.type%3Danswer%29%5D.target.author.badge%5B%3F%28type%3Dbest_answerer%29%5D.topics%3Bdata%5B%3F%28target.type%3Darticle%29%5D.target.annotation_detail%2Ccontent%2Chermes_label%2Cis_labeled%2Cauthor.badge%5B%3F%28type%3Dbest_answerer%29%5D.topics%3Bdata%5B%3F%28target.type%3Dquestion%29%5D.target.annotation_detail%2Ccomment_count%3B&limit=5'
        l = 'https://www.zhihu.com/topic/'+id+'/hot'
        work.put_nowait((url, l))
    def crawler():
        while not work.empty():
            # 提取url
            url,l = work.get_nowait()
            headers = {'user-agent':ua_pond()}
            # 提取话题名称
            rl = requests.get(l,headers=headers,proxies=proxy_pond())
            res = BeautifulSoup(rl.text, 'html.parser')
            if res.find(class_='ContentItem-title') != None:
                topic = res.find(class_='ContentItem-title').text
            else:
                topic = '未找到'
            # 提取话题里的问题
            r = requests.get(url,headers=headers,proxies=proxy_pond())
            soup = r.json()
            for i in soup['data']:
                try:
                    title = i['target']['question']['title']
                    question = 'question/' + re.findall('s/.*?(\d.*\d)', i['target']['question']['url'])[0]
                    answer = '/answer/' + re.findall('s/.*?(\d.*\d)', i['target']['url'])[0]
                    # 构造回答的url
                    url = 'https://www.zhihu.com/' + question + answer
                    # 将标题和url存储到数据库
                    add_data([topic,title,url])
                # 有的标题存放路径不同
                except:
                    title = i['target']['title']
                    try:
                        id = i['target']['id']
                        url = 'https://zhuanlan.zhihu.com/p/' + str(id)
                    except:
                        url = '无'
                    add_data([topic,title,url])
            while soup['paging']['is_end'] == False:
                url = soup['paging']['next']
                res = requests.get(url,headers=headers,proxies=proxy_pond())
                soup = res.json()
                for i in soup['data']:
                    try:
                        title = i['target']['question']['title']
                        question = 'question/' + re.findall('s/.*?(\d.*\d)', i['target']['question']['url'])[0]
                        answer = '/answer/' + re.findall('s/.*?(\d.*\d)', i['target']['url'])[0]
                        # 构造回答的url
                        url = 'https://www.zhihu.com/' + question + answer
                        # 将标题和url存储到数据库
                        add_data([topic,title,url])
                    except:
                        title = i['target']['title']
                        try:
                            id = i['target']['id']
                            url = 'https://zhuanlan.zhihu.com/p/' + str(id)
                        except:
                            url = '无'
                        add_data([topic,title,url])
    # 创建5只爬虫
    tasks_list = []
    for i in range(5):
        task = gevent.spawn(crawler)
        tasks_list.append(task)

    # 开始执行队列任务
    gevent.joinall(tasks_list)

# 主函数
def main():
    start = time.time()
    my_url()
    produce()
    consumer()
    end = time.time()
    print(end-start)

if __name__ == '__main__':
    main()


